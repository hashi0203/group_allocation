# -*- coding: utf-8 -*-
import csv
import random
import collections
import copy
import openpyxl
from openpyxl.styles import PatternFill

# 人数調整
def arrange(memlist,n):
    len_initmemlist = len(memlist)
    if n <= 3:
        for i in range(len_memlist):
            if [memlist[i][0], memlist[i][2]] in early_leave:
                del memlist[i]
                break
        added = (n-(len_initmemlist-num_early_leave)%n)%n
        print('number of ECs: ' + str(added))
        for i in range(added):
            memlist += [[17, len_initmemlist+i, 'EC']]
    else:
        added = (n-len_initmemlist%n)%n
        print('number of ECs: ' + str(added))
        for i in range(added):
            memlist += [[17, len_initmemlist+i, 'EC']]
    return memlist

# n人ずつで区切る
def mk_group(memlist,n):
    len_memlist = len(memlist)
    group = []
    mem = []
    for i in range(len_memlist):
        mem += [memlist[i]]
        if i%n == n-1:
            mem.sort(key=lambda x:x[0])
            group += [[mem]]
            mem = []
    group[len_memlist//n-1].sort(key=lambda x:x[0])
    return group


# 制約を元に重みづけ
def mk_constraints_4(group):
    for i in range(len_group):
        member = []
        for j in range(4):
            member += [group[i][0][j][0]]
        c = collections.Counter(member)
        group[i] += [[c[15],c[16],c[17],c[18],c[19]]]
        if c.most_common()[0][1] >= 3:
            group[i] += [[c.most_common()[0][0]-15]] + [[5]]
        elif c[15]//2 == 2:
            group[i] += [[0]] + [[5]]
        elif c[15]//2 + c[16]//2 + c[17]//2 + c[18]//2 + c[19]//2 == 2:
            group[i] += [[c.most_common()[0][0]-15]] + [[5]]
        elif c.most_common()[0][1] == 2:
            group[i] += [[c.most_common()[0][0]-15]] + [[2]]
        else:
            group[i] += [[c.most_common()[0][0]-15]] + [[0]]
    # 重みの高いものから順に並べる
    group.sort(key=lambda x:x[3],reverse=True)

# swap時の重みの書き換え
def chng_constraints_4(group,k,l):
    for i in [k,l]:
        member = []
        for j in range(4):
            member += [group[i][0][j][0]]
        c = collections.Counter(member)
        group[i][1] = [c[15],c[16],c[17],c[18],c[19]]
        if c.most_common()[0][1] >= 3:
            group[i][2] = [c.most_common()[0][0]-15]
            group[i][3] = [5]
        elif c[15]//2 == 2:
            group[i][2] = [0]
            group[i][3] = [5]
        elif c[15]//2 + c[16]//2 + c[17]//2 + c[18]//2 + c[19]//2 == 2:
            group[i][2] = [c.most_common()[0][0]-15]
            group[i][3] = [4]
        elif c.most_common()[0][1] == 2:
            group[i][2] = [c.most_common()[0][0]-15]
            group[i][3] = [2]
        else:
            group[i][2] = [c.most_common()[0][0]-15]
            group[i][3] = [0]

# 制約を元に重みづけ
def mk_constraints_3(group):
    for i in range(len_group):
        member = []
        for j in range(3):
            member += [group[i][0][j][0]]
        c = collections.Counter(member)
        group[i] += [[c[15],c[16],c[17],c[18],c[19]]]
        if c.most_common()[0][1] >= 3:
            group[i] += [[c.most_common()[0][0]-15]] + [[5]]
        elif c[15]//2 == 2:
            group[i] += [[0]] + [[5]]
        elif c.most_common()[0][1] == 2:
            group[i] += [[c.most_common()[0][0]-15]] + [[4]]
        else:
            group[i] += [[c.most_common()[0][0]-15]] + [[0]]
    # 重みの高いものから順に並べる
    group.sort(key=lambda x:x[3],reverse=True)

# swap時の重みの書き換え
def chng_constraints_3(group,k,l):
    for i in [k,l]:
        member = []
        for j in range(3):
            member += [group[i][0][j][0]]
        c = collections.Counter(member)
        group[i][1] = [c[15],c[16],c[17],c[18],c[19]]
        if c.most_common()[0][1] >= 3:
            group[i][2] = [c.most_common()[0][0]-15]
            group[i][3] = [5]
        elif c[15]//2 == 2:
            group[i][2] = [0]
            group[i][3] = [5]
        elif c.most_common()[0][1] == 2:
            group[i][2] = [c.most_common()[0][0]-15]
            group[i][3] = [4]
        else:
            group[i][2] = [c.most_common()[0][0]-15]
            group[i][3] = [0]

# 制約を元に重みづけ
def mk_constraints_2(group):
    for i in range(len_group):
        member = []
        for j in range(2):
            member += [group[i][0][j][0]]
        c = collections.Counter(member)
        group[i] += [[c[15],c[16],c[17],c[18],c[19]]]
        if c[19] == 1 and (c[17] == 1 or c[18] == 1):
            group[i] += [[c.most_common()[0][0]-15]] + [[0]]
        elif c[19] == 2:
            group[i] += [[c.most_common()[0][0]-15]] + [[4]]
        elif c.most_common()[0][1] == 2:
            group[i] += [[c.most_common()[0][0]-15]] + [[5]]
        elif c[15] == 1 and c[19] == 1:
            group[i] += [[c.most_common()[0][0]-15]] + [[4]]
        elif c[16] == 1 and c[19] == 1:
            group[i] += [[c.most_common()[0][0]-15]] + [[4]]
        elif c[18] == 1 and c[15] == 1:
            group[i] += [[c.most_common()[0][0]-15]] + [[2]]
        else:
            group[i] += [[c.most_common()[0][0]-15]] + [[1]]
    # 重みの高いものから順に並べる
    group.sort(key=lambda x:x[3],reverse=True)

# swap時の重みの書き換え
def chng_constraints_2(group,k,l):
    for i in [k,l]:
        member = []
        for j in range(2):
            member += [group[i][0][j][0]]
        c = collections.Counter(member)
        group[i][1] = [c[15],c[16],c[17],c[18],c[19]]
        if c[19] == 1 and (c[17] == 1 or c[18] == 1):
            group[i][2] = [c.most_common()[0][0]-15]
            group[i][3] = [0]
        elif c[19] == 2:
            group[i][2] = [c.most_common()[0][0]-15]
            group[i][3] = [4]
        elif c.most_common()[0][1] == 2:
            group[i][2] = [c.most_common()[0][0]-15]
            group[i][3] = [5]
        elif c[15] == 1 and c[19] == 1:
            group[i][2] = [c.most_common()[0][0]-15]
            group[i][3] = [4]
        elif c[16] == 1 and c[19] == 1:
            group[i][2] = [c.most_common()[0][0]-15]
            group[i][3] = [4]
        elif c[18] == 1 and c[15] == 1:
            group[i][2] = [c.most_common()[0][0]-15]
            group[i][3] = [2]
        else:
            group[i][2] = [c.most_common()[0][0]-15]
            group[i][3] = [1]

# 条件を満たすようにswapする
def swap(group,i,j,fl,n):
    if n == 2:
        a = 0
        b = 1
    else:
        a = fl
        while group[i][0][a][0] != group[i][2][0]+15:
            if a == len(group[j][0]) - 1:
                break
            a += 1
        b = len(group[j][0]) - 1
        while group[j][0][b][0] != group[j][2][0]+15:
            if b == fl:
                break
            b -= 1
    group[i][0][a], group[j][0][b] = group[j][0][b], group[i][0][a]
    if fl == 0:
        group[i][0].sort(key=lambda x:x[0])
        group[j][0].sort(key=lambda x:x[0])
    if n == 4:
        chng_constraints_4(group,i,j)
    elif n == 3:
        chng_constraints_3(group,i,j)
    elif n == 2:
        chng_constraints_2(group,i,j)



# 制約を満たすグループ分けを作る
def prv_constraints(group,fl,n):
    if n == 2:
        i = 0
        count = 0
        while i < len_group and count < 10000:
            count += 0
            # 重みが5以上なら
            if group[i][3][0] >= 5:
                j = i+1
                while group[j][3][0] != 0 and j < len_group-1:
                    j += 1
                swap(group,i,j,fl,n)
                # 同じ状況でループすることがあるので、シャッフルして並び順を変える
                if count%30 == 0:
                    random.shuffle(group)
                # 重みの高いものから順に並べる
                group.sort(key=lambda x:x[3],reverse=True)
                i = 0
            else:
                i += 1
        if count == 10000:
            return 0
        i = 0
        count = 0
        while i < len_group and count < 10000:
            count += 1
            # 重みが4以上なら
            if group[i][3][0] >= 4:
                j = i+1
                while (group[j][3][0] != 1 or (count < 5000 and group[j][3][0] != 2)) and j < len_group-1:
                    j += 1
                swap(group,i,j,fl,n)
                # 同じ状況でループすることがあるので、シャッフルして並び順を変える
                if count%30 == 0:
                    random.shuffle(group)
                # 重みの高いものから順に並べる
                group.sort(key=lambda x:x[3],reverse=True)
                i = 0
            else:
                i += 1
        i = 0
        count = 0
        while i < len_group and count < 10000:
            count += 1
            # 重みが2以上なら
            if group[i][3][0] >= 2:
                j = i+1
                while group[j][3][0] != 1 and j < len_group-1:
                    j += 1
                swap(group,i,j,fl,n)
                # 同じ状況でループすることがあるので、シャッフルして並び順を変える
                if count%30 == 0:
                    random.shuffle(group)
                # 重みの高いものから順に並べる
                group.sort(key=lambda x:x[3],reverse=True)
                i = 0
            else:
                i += 1
    else:
        i = 0
        count = 0
        while i < len_group and count < 5000:
            count += 1
            # 重みが5以上なら
            if group[i][3][0] >= 5:
                j = i+1
                while (group[j][1][group[i][2][0]] != 0 or group[j][1][0] >= 1) and j < len_group-1:
                    j += 1
                swap(group,i,j,fl,n)
                # 同じ状況でループすることがあるので、シャッフルして並び順を変える
                if count%30 == 0:
                    random.shuffle(group)
                # 重みの高いものから順に並べる
                group.sort(key=lambda x:x[3],reverse=True)
                i = 0
            else:
                i += 1
        if count == 5000:
            return 0
        # 重みが4以上のものをできるだけ減らす (5000回繰り返して無理ならあきらめていい)
        i = 0
        count = 0
        while i < len_group and count < 5000:
            count += 1
            # 重みが4以上なら
            if group[i][3][0] >= 4:
                j = i+1
                while (group[j][1][group[i][2][0]] != 0 or group[j][1][0] >= 1) and j < len_group-1:
                    j += 1
                swap(group,i,j,fl,n)
                # 同じ状況でループすることがあるので、シャッフルして並び順を変える
                if count%30 == 0:
                    random.shuffle(group)
                # 重みの高いものから順に並べる
                group.sort(key=lambda x:x[3],reverse=True)
                i = 0
            else:
                i += 1
        # 重みが2以上のものをできるだけ減らす (5000回繰り返して無理ならあきらめていい)
        i = 0
        count = 0
        while i < len_group and count < 5000:
            count += 1
            # 重みが2以上なら
            if group[i][3][0] >= 2:
                j = i+1
                while (group[j][1][group[i][2][0]] != 0 or group[j][1][0] >= 1) and j < len_group-1:
                    j += 1
                swap(group,i,j,fl,n)
                # 同じ状況でループすることがあるので、シャッフルして並び順を変える
                if count%30 == 0:
                    random.shuffle(group)
                # 重みの高いものから順に並べる
                group.sort(key=lambda x:x[3],reverse=True)
                i = 0
            else:
                i += 1
    return 1

# member1[j]とmember2[j]をswapしても制約を満たすなら1を返す．
def check_mem(member1,member2,j,m,n):
    # 現在確認中のグループが入れ替え後全員speakerを経験していたらダメ
    if n != 2:
        flag = 0
        if speaker[member2[m][1]] == 1:
            for a in range(n):
                if a != j and speaker[member1[a][1]] == 0:
                    flag = 1
            if flag == 0:
                return 0
    # 入れ替えた先で同じグループになった人がいればダメ
    for a in range(n):
        if a != j and mem_in_same_group[member1[a][1]][member2[m][1]] == 1:
            return 0
    for b in range(n):
        if b != m and mem_in_same_group[member1[j][1]][member2[b][1]] == 1:
            return 0
    return 1

def can_swap(group,i,j,k,n):
    # group i 内の j と k が同じグループになったことがあるとき
    if j!= k and mem_in_same_group[group[i][0][j][1]][group[i][0][k][1]] == 1:
        l = (i + 1) % len_group
        grade_j = group[i][0][j][0]
        grade_k = group[i][0][k][0]
        # 一周回るまで繰り替えす
        while l != i:
            for m in range(n):
                if l > i or m > 0:
                    # 同じ学年の人を入れ替える
                    if group[l][0][m][0] == grade_j:
                        if check_mem(group[i][0],group[l][0],j,m,n) == 1:
                            group[i][0][j], group[l][0][m] = group[l][0][m], group[i][0][j]
                            return 1
                    if group[l][0][m][0] == grade_k:
                        if check_mem(group[i][0],group[l][0],k,m,n) == 1:
                            group[i][0][k], group[l][0][m] = group[l][0][m], group[i][0][k]
                            return 1
            l = (l + 1) % len_group
        return 0

# 同じ人と2回同じグループにならないように調整する
def swap_same_mem(group,n):
    for i in range(len_group):
        # print(i)
        for j in range(n):
            for k in range(n):
                if can_swap(group,i,j,k,n) == 0:
                    return 0
        if n != 2:
            # speakerになったことない人の中からrandomにspeakerを選び、各組の一番前にする
            tmpmem = list(filter(lambda n:speaker[n[1]]==0, group[i][0]))
            if len(tmpmem) == 0:
                return 0
            random.shuffle(tmpmem)
            for l in range(n):
                if group[i][0][l] == tmpmem[0]:
                    for m in range(l):
                        group[i][0][l-m], group[i][0][l-m-1] = group[i][0][l-m-1], group[i][0][l-m]
                    # if speaker[tmpmem[0][1]] == 1:
                    #     print(tmpmem[0][1])
                    speaker[tmpmem[0][1]] = 1
                    break
        else:
            # 2の時は上級生が先に聞く側になる
            group[i][0][0], group[i][0][1] = group[i][0][1], group[i][0][0]

    return 1
    

# 同じグループになったことがある人を更新
def renew_same(group,table,n):
    for i in range(len_group):
        for j in range(n):
            for k in range(n):
                # if j != k and table[group[i][0][j][1]][group[i][0][k][1]] == 1:
                #     print(group[i][0][j][2] + "," + group[i][0][k][2] + "," + str(jj))
                table[group[i][0][j][1]][group[i][0][k][1]] = 1

def group_of_4(memlist,len_memlist,len_group):
    global speaker
    for i in range(4):
        save_speaker = copy.copy(speaker)
        random.shuffle(memlist)
        if i == 3:
            mem_nonsp = list(filter(lambda n:speaker[n[1]]==0, memlist))
            group1 = mk_group(mem_nonsp,1)
            mem_sp = list(filter(lambda n:speaker[n[1]]==1, memlist))
            group2 = mk_group(mem_sp,3)
            group = []
            for g12 in range(len(group1)):
                group += [[group1[g12][0] + group2[g12][0]]]
            mk_constraints_4(group)
            if prv_constraints(group,1,4) == 0:
                return 0
        else:
            group = mk_group(memlist,4)
            mk_constraints_4(group)
            if prv_constraints(group,0,4) == 0:
                return 0
        j = 0
        while j < 20 and swap_same_mem(group,4) == 0:
            j += 1
            speaker = copy.copy(save_speaker)
            random.shuffle(memlist)
            if i == 3:
                mem_nonsp = list(filter(lambda n:speaker[n[1]]==0, memlist))
                group1 = mk_group(mem_nonsp,1)
                mem_sp = list(filter(lambda n:speaker[n[1]]==1, memlist))
                group2 = mk_group(mem_sp,3)
                group = []
                for g12 in range(len(group1)):
                    group += [[group1[g12][0] + group2[g12][0]]]
                mk_constraints_4(group)
                if prv_constraints(group,1,4) == 0:
                    return 0
            else:
                group = mk_group(memlist,4)
                mk_constraints_4(group)
                if prv_constraints(group,0,4) == 0:
                    return 0
        if j == 20:
            return 0
        renew_same(group,mem_in_same_group,4)

        # output.xlsxに出力
        output = []
        for k in range(len_group):
            output += group[k][0]

        sheet = book['4-term' + str(i+1)]

        for k in range(150):
            if k < 68:
                if k % 4 == 0:
                    sheet.cell(row=k + 3,column=2).fill = fill
                    sheet.cell(row=k + 3,column=3).fill = fill
                if output[k][0] == 15:
                    sheet.cell(row=k + 3,column=2).value = 20
                else:
                    sheet.cell(row=k + 3,column=2).value = output[k][0]
                sheet.cell(row=k + 3,column=3).value = output[k][2]
            elif k < len_memlist:
                if k % 4 == 0:
                    sheet.cell(row=k - 65,column=5).fill = fill
                    sheet.cell(row=k - 65,column=6).fill = fill
                if output[k][0] == 15:
                    sheet.cell(row=k - 65,column=5).value = 20
                else:
                    sheet.cell(row=k - 65,column=5).value = output[k][0]
                sheet.cell(row=k - 65,column=6).value = output[k][2]
            else:
                sheet.cell(row=k - 65,column=5).fill = nofill
                sheet.cell(row=k - 65,column=6).fill = nofill
                sheet.cell(row=k - 65,column=5).value = ''
                sheet.cell(row=k - 65,column=6).value = ''

        # group_of_4.csvに出力
        result = []
        for k in range(len_group):
            result += [group[k][0] + [[group[k][3][0]]]]

        if i == 0:
            with open('group_of_4.csv', 'w') as fo:
                writer = csv.writer(fo, lineterminator='\n')
                writer.writerow(['term'+str(i+1)])
                writer.writerows(result)
                # writer.writerows(mem_in_same_group)
        else:
            with open('group_of_4.csv', 'a') as fo:
                writer = csv.writer(fo, lineterminator='\n')
                writer.writerow(['term'+str(i+1)])
                writer.writerows(result)
                # debug用
                # debug = []
                # for k in range(len_group):
                #     debug += [[group[k][0][0][1]]] + [[group[k][0][1][1]]] + [[group[k][0][2][1]]] + [[group[k][0][3][1]]]
                # debug.sort(key=lambda x:x)
                # writer.writerows(debug)
                # writer.writerows(mem_in_same_group)
    return 1

def group_of_3(memlist,len_memlist,len_group):
    global speaker
    for i in range(3):
        save_speaker = copy.copy(speaker)
        random.shuffle(memlist)
        if i == 2:
            mem_nonsp = list(filter(lambda n:speaker[n[1]]==0, memlist))
            group1 = mk_group(mem_nonsp,1)
            mem_sp = list(filter(lambda n:speaker[n[1]]==1, memlist))
            group2 = mk_group(mem_sp,2)
            group = []
            for g12 in range(len(group1)):
                group += [[group1[g12][0] + group2[g12][0]]]
            mk_constraints_3(group)
            if prv_constraints(group,1,3) == 0:
                return 0
        else:
            group = mk_group(memlist,3)
            mk_constraints_3(group)
            if prv_constraints(group,0,3) == 0:
                return 0
        j = 0
        while j < 30 and swap_same_mem(group,3) == 0:
            j += 1
            speaker = copy.copy(save_speaker)
            random.shuffle(memlist)
            if i == 2:
                mem_nonsp = list(filter(lambda n:speaker[n[1]]==0, memlist))
                group1 = mk_group(mem_nonsp,1)
                mem_sp = list(filter(lambda n:speaker[n[1]]==1, memlist))
                group2 = mk_group(mem_sp,2)
                group = []
                for g12 in range(len(group1)):
                    group += [[group1[g12][0] + group2[g12][0]]]
                mk_constraints_3(group)
                if prv_constraints(group,1,3) == 0:
                    return 0
            else:
                group = mk_group(memlist,3)
                mk_constraints_3(group)
                if prv_constraints(group,0,3) == 0:
                    return 0
        if j == 30:
            return 0
        renew_same(group,mem_in_same_group,3)

        # output.xlsxに出力
        output = []
        for k in range(len_group):
            output += group[k][0]

        sheet = book['3-term' + str(i+1)]

        for k in range(150):
            if k < 66:
                if k % 3 == 0:
                    sheet.cell(row=k + 3,column=2).fill = fill
                    sheet.cell(row=k + 3,column=3).fill = fill
                if output[k][0] == 15:
                    sheet.cell(row=k + 3,column=2).value = 20
                else:
                    sheet.cell(row=k + 3,column=2).value = output[k][0]
                sheet.cell(row=k + 3,column=3).value = output[k][2]
            elif k < len_memlist:
                if k % 3 == 0:
                    sheet.cell(row=k - 63,column=5).fill = fill
                    sheet.cell(row=k - 63,column=6).fill = fill
                if output[k][0] == 15:
                    sheet.cell(row=k - 63,column=5).value = 20
                else:
                    sheet.cell(row=k - 63,column=5).value = output[k][0]
                sheet.cell(row=k - 63,column=6).value = output[k][2]
            else:
                sheet.cell(row=k - 63,column=5).fill = nofill
                sheet.cell(row=k - 63,column=6).fill = nofill
                sheet.cell(row=k - 63,column=5).value = ''
                sheet.cell(row=k - 63,column=6).value = ''

        # group_of_3.csvに出力
        result = []
        for k in range(len_group):
            result += [group[k][0] + [[group[k][3][0]]]]

        if i == 0:
            with open('group_of_3.csv', 'w') as fo:
                writer = csv.writer(fo, lineterminator='\n')
                writer.writerow(['term'+str(i+5)])
                writer.writerows(result)
                # writer.writerows(mem_in_same_group)
        else:
            with open('group_of_3.csv', 'a') as fo:
                writer = csv.writer(fo, lineterminator='\n')
                writer.writerow(['term'+str(i+5)])
                writer.writerows(result)
                # debug用
                # debug = []
                # for k in range(len_group):
                #     debug += [[group[k][0][0][1]]] + [[group[k][0][1][1]]] + [[group[k][0][2][1]]]
                # debug.sort(key=lambda x:x)
                # writer.writerows(debug)
                # writer.writerows(mem_in_same_group)
    return 1

def group_of_2(memlist,len_memlist,len_group):
    random.shuffle(memlist)
    group = mk_group(memlist,2)
    mk_constraints_2(group)
    if prv_constraints(group,0,2) == 0:
        return 0

    j = 0
    while j < 20 and swap_same_mem(group,2) == 0:
        j += 1
        random.shuffle(memlist)
        group = mk_group(memlist,2)
        mk_constraints_2(group)
        if prv_constraints(group,0,2) == 0:
            return 0
    if j == 20:
        return 0
    renew_same(group,mem_in_same_group,2)

    # output.xlsxに出力
    output = []
    for k in range(len_group):
        output += group[k][0]

    sheet = book['2']

    for k in range(150):
        if k < 68:
            if k % 2 == 0:
                sheet.cell(row=k + 3,column=2).fill = fill
                sheet.cell(row=k + 3,column=3).fill = fill
            if output[k][0] == 15:
                    sheet.cell(row=k + 3,column=2).value = 20
            else:
                sheet.cell(row=k + 3,column=2).value = output[k][0]
            sheet.cell(row=k + 3,column=3).value = output[k][2]
        elif k < len_memlist:
            if k % 2 == 0:
                sheet.cell(row=k - 65,column=5).fill = fill
                sheet.cell(row=k - 65,column=6).fill = fill
            if output[k][0] == 15:
                    sheet.cell(row=k - 65,column=5).value = 20
            else:
                sheet.cell(row=k - 65,column=5).value = output[k][0]
            sheet.cell(row=k - 65,column=6).value = output[k][2]
        else:
            sheet.cell(row=k - 65,column=5).fill = nofill
            sheet.cell(row=k - 65,column=6).fill = nofill
            sheet.cell(row=k - 65,column=5).value = ''
            sheet.cell(row=k - 65,column=6).value = ''

    # group_of_2.csvに出力
    result = []
    for k in range(len_group):
        result += [group[k][0] + [[group[k][3][0]]]]
    with open('group_of_2.csv', 'w') as fo:
        writer = csv.writer(fo, lineterminator='\n')
        writer.writerow(['term'+str(8)])
        writer.writerows(result)
        # debug用
        # debug = []
        # for k in range(len_group):
        #     debug += [[group[k][0][0][1]]] + [[group[k][0][1][1]]]
        # debug.sort(key=lambda x:x)
        # writer.writerows(debug)
        # writer.writerows(mem_in_same_group)
    return 1


# メイン関数
fill = PatternFill(patternType='solid', fgColor='d3d3d3')
nofill = PatternFill(patternType='solid', fgColor='ffffff')
# 欠席者情報読み込み
absentee = []
with open('absentee.csv','r') as fi:
    reader = csv.reader(fi)
    for row in reader:
        absentee += [[int(row[0]), row[1]]]
# 早退者情報読み込み
early_leave = []
with open('early_leave.csv','r') as fi:
    reader = csv.reader(fi)
    for row in reader:
        early_leave += [[int(row[0]), row[1]]]
num_early_leave = len(early_leave)
# import data
initmemlist = []
serial = 0
with open('memlist.csv','r') as fi:
    reader = csv.reader(fi)
    print('Absentee')
    for row in reader:
        if [int(row[0]), row[1]] in absentee:
            print(row[1])
        else:
            initmemlist += [[int(row[0]), serial, row[1]]]
            serial += 1
len_initmemlist = len(initmemlist)+3
# 同じにグループになったことのある人の表(なったことがあれば1なければ0)最大で(人数+3)人になる
mem_in_same_group = [[0] * len_initmemlist for i in range(len_initmemlist)]
# 話し手になったことのある人の表(なったことがあれば1なければ0)最大で(人数+3)人になる
speaker = [0 for i in range(len_initmemlist)]

book = openpyxl.load_workbook('output.xlsx')

# 4人グループの作成
print('\ngroup_of_4 start')
memlist = arrange(copy.copy(initmemlist),4)
len_memlist = len(memlist)
len_group = len_memlist//4
for retry in range(20):
    mem_in_same_group = [[0] * len_initmemlist for retry in range(len_initmemlist)]
    speaker = [0 for i in range(len_initmemlist)]
    if group_of_4(memlist,len_memlist,len_group) == 1:
        # print('話し手になった回数')
        # print(speaker)
        print("group_of_4 done")
        break
    if retry == 19:
        print("no answer")
# 3人グループの作成
print('\ngroup_of_3 start')
memlist = arrange(copy.copy(initmemlist),3)
len_memlist = len(memlist)
len_group = len_memlist//3
save_mem_in_same_group = copy.deepcopy(mem_in_same_group)
for retry in range(20):
    mem_in_same_group = copy.deepcopy(save_mem_in_same_group)
    speaker = [0 for i in range(len_initmemlist)]
    if group_of_3(memlist,len_memlist,len_group) == 1:
        # print('話し手になった回数')
        # print(speaker)
        print("group_of_3 done")
        break
    if retry == 19:
        print("no answer")
# 2人グループの作成
print('\ngroup_of_2 start')
memlist = arrange(copy.copy(initmemlist),2)
len_memlist = len(memlist)
len_group = len_memlist//2
save_mem_in_same_group = copy.deepcopy(mem_in_same_group)
for retry in range(20):
    mem_in_same_group = copy.deepcopy(save_mem_in_same_group)
    if group_of_2(memlist,len_memlist,len_group) == 1:
        print('Number of people who were in the same group')
        same_group = []
        for r in range(len(mem_in_same_group)):
            same_group += [sum(c for c in mem_in_same_group[r])]
        print(same_group)
        print("group_of_2 done")
        break
    if retry == 19:
        print("no answer")

book.save('output.xlsx')
print("finish")
